from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
import openpyxl
import time
import datetime
import os
import shutil

chrome_options = webdriver.ChromeOptions()
'''chrome_options.add_argument('--headless')
chrome_options.add_argument('--disable-gpu')'''

driver = webdriver.Chrome(service=ChromeService(), options=chrome_options)
driver.get('https://codis.cwa.gov.tw/StationData')
driver.set_window_size(1660, 945)
driver.implicitly_wait(5)

time.sleep(2)
#改成清單模式，並選擇只看北部地區
driver.find_element(By.CSS_SELECTOR, '#switch_display > button:nth-child(2)').click()
driver.find_element(By.CSS_SELECTOR, '#cwb').click()
driver.find_element(By.CSS_SELECTOR, '#auto_C1').click()
driver.find_element(By.CSS_SELECTOR, '#station_filter > div > section > ul > li:nth-child(3) > div > div.col-7.px-0 > select > option:nth-child(2)').click()

#選擇站點
times = driver.find_element(By.CSS_SELECTOR, '#station_count').text
for nowtimes in range(1, int(times)+1): #用 range 開始的數字來調整從哪個測站開始
    name = driver.find_element(By.CSS_SELECTOR, '#station_table > table > tbody > tr:nth-child({}) > td:nth-child(1) > div'.format(nowtimes)).text
    if('雷達站' in name):
        print("雷達站跳過")
        continue
    print("目前測站{} NO.{}，還有{}個測站".format(name, nowtimes, str(int(times)-nowtimes)))
    number = driver.find_element(By.CSS_SELECTOR, '#station_table > table > tbody > tr:nth-child({}) > td:nth-child(2) > div'.format(nowtimes)).text
    driver.find_element(By.CSS_SELECTOR, '#station_table > table > tbody > tr:nth-child({}) > td:nth-child(10) > div'.format(nowtimes)).click()

    #提取站點資訊
    place = driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div.lightbox-tool-station > div.lightbox-tool-station-info > div.lightbox-tool-station-subinfo > div:nth-child(1)').text.split('：')[1]
    longitude = driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div.lightbox-tool-station > div.lightbox-tool-station-info > div.lightbox-tool-station-subinfo > div:nth-child(2)').text.split('：')[1].split(',')[0][:-2]
    latitude = driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div.lightbox-tool-station > div.lightbox-tool-station-info > div.lightbox-tool-station-subinfo > div:nth-child(2)').text.split('：')[1].split(',')[1][:-2]
    altitude = driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div.lightbox-tool-station > div.lightbox-tool-station-info > div.lightbox-tool-station-subinfo > div:nth-child(3)').text.split('：')[1]
    date = driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div.lightbox-tool-station > div.lightbox-tool-station-info > div.lightbox-tool-station-subinfo > div:nth-child(4)').text[5:15]
    date = list(map(int, date.split('-')))
    
    '''
    if(date[0] < 2021):
        date = [2021, 1, 1]
    else:
        pass
    '''
    #要客製化某個測站開始月份時用的
    if(nowtimes == 22):
        date = [2022, 11, 22]
    elif(date[0] < 2021):
        date = [2021, 1, 1]
    else:
        pass

    #將資料輸入進excel
    workbook = openpyxl.Workbook()
    sheet = workbook.worksheets[0]
    chtitle = ['站名', '縣市', '經度', '緯度', '海拔高度']
    entitle = ["Name", "Place", "Longitude", "Latitude", "Altitude"]
    for num in range(len(chtitle)):
        sheet.cell(row=1, column=num+1).value = chtitle[num]
        sheet.cell(row=2, column=num+1).value = entitle[num]
    sheet.cell(row = 3, column = 1).value = name
    sheet.cell(row = 3, column = 2).value = place
    sheet.cell(row = 3, column = 3).value = longitude
    sheet.cell(row = 3, column = 4).value = latitude
    sheet.cell(row = 3, column = 5).value = altitude
    try:
        os.mkdir('{}({})'.format(name, number))
    except:
        pass
    workbook.save(os.path.join(os.getcwd(), '{}({})'.format(name, number), '{}({}).xlsx'.format(name, number)))
    workbook.close()
    
    #修改觀測年月
    driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div:nth-child(5) > div.lightbox-tool-type-ctrl > div.lightbox-tool-type-ctrl-form > label > div > div.vdatetime > input').click()
    time.sleep(1)

    #年
    nyear = driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div:nth-child(5) > div.lightbox-tool-type-ctrl > div.lightbox-tool-type-ctrl-form > label > div > div.vdatetime > div > div.vdatetime-popup > div.vdatetime-popup__header > div.vdatetime-popup__year').text
    driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div:nth-child(5) > div.lightbox-tool-type-ctrl > div.lightbox-tool-type-ctrl-form > label > div > div.vdatetime > div > div.vdatetime-popup > div.vdatetime-popup__header > div.vdatetime-popup__year').click()
    driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div:nth-child(5) > div.lightbox-tool-type-ctrl > div.lightbox-tool-type-ctrl-form > label > div > div.vdatetime > div > div.vdatetime-popup > div.vdatetime-popup__body > div > div > div:nth-child({})'.format(100-(int(nyear)-date[0]-1))).click()

    now = driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div:nth-child(5) > div.lightbox-tool-type-ctrl > div.lightbox-tool-type-ctrl-form > label > div > div.vdatetime > div > div.vdatetime-popup > div.vdatetime-popup__header > div.vdatetime-popup__date').text
    #月
    nmin = now.split('月')[0]
    driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div:nth-child(5) > div.lightbox-tool-type-ctrl > div.lightbox-tool-type-ctrl-form > label > div > div.vdatetime > div > div.vdatetime-popup > div.vdatetime-popup__header > div.vdatetime-popup__date').click()
    if(date[1] == nmin):
        driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div:nth-child(5) > div.lightbox-tool-type-ctrl > div.lightbox-tool-type-ctrl-form > label > div > div.vdatetime > div > div.vdatetime-popup > div.vdatetime-popup__body > div > div > div.vdatetime-month-picker__item.vdatetime-month-picker__item--selected').click()
    else:
        driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div:nth-child(5) > div.lightbox-tool-type-ctrl > div.lightbox-tool-type-ctrl-form > label > div > div.vdatetime > div > div.vdatetime-popup > div.vdatetime-popup__body > div > div > div:nth-child({})'.format(date[1])).click()

    #日
    nday = now.split('月')[1].replace('日', '')
    if(date[2] == nday):
        driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div:nth-child(5) > div.lightbox-tool-type-ctrl > div.lightbox-tool-type-ctrl-form > label > div > div.vdatetime > div > div.vdatetime-popup > div.vdatetime-popup__body > div > div.vdatetime-calendar__month > div.vdatetime-calendar__month__day.vdatetime-calendar__month__day--selected').click()
    else:
        for i in range(8, 43):
            tmp = driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div:nth-child(5) > div.lightbox-tool-type-ctrl > div.lightbox-tool-type-ctrl-form > label > div > div.vdatetime > div > div.vdatetime-popup > div.vdatetime-popup__body > div > div.vdatetime-calendar__month > div:nth-child({})'.format(i))
            if(tmp.text == str(date[2])):
                tmp.click()
                break

    driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div:nth-child(5) > div.lightbox-tool-type-ctrl > div.lightbox-tool-type-ctrl-form > label > div > div.vdatetime > div > div.vdatetime-popup > div.vdatetime-popup__actions > div.vdatetime-popup__actions__button.vdatetime-popup__actions__button--confirm').click()
    
    #計算該觀測站共有幾個月的資料須爬取
    if(date == [2023, 12, 31]):
        daynum = 1
    else:
        daynum = datetime.datetime(2023, 12, 31) - datetime.datetime(date[0], date[1], date[2])
        daynum = int(str(daynum).split(' ')[0])+1
    print('還有 {} 天'.format(daynum))

    time.sleep(1)
    title_list = driver.find_elements(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div:nth-child(8) > div.lightbox-tool-type-ctrl > div.lightbox-tool-type-ctrl-form > div:nth-child(2) > label.w-100 > select > option')

    for _ in range(daynum):
        driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div:nth-child(5) > div.lightbox-tool-type-ctrl > div.lightbox-tool-type-ctrl-btn-group > div').click()
        driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div:nth-child(5) > div.lightbox-tool-type-ctrl > div.lightbox-tool-type-ctrl-form > label > div > div:nth-child(5)').click()
        time.sleep(1)
        if((_+1) % 30 == 0):
            filelist = os.listdir('C:\\Users\\GIGABYTE\\Downloads')
            for filename in filelist:
                if(number in filename):
                    shutil.move('C:\\Users\\GIGABYTE\\Downloads\\' + filename, "C:\\Users\\GIGABYTE\\桌面\\VScode_Code\\天氣預測資料\\自動雨量站\\" + '{}({})'.format(name, number))
            ndate = datetime.datetime(date[0], date[1], date[2]) + datetime.timedelta(days=_)
            print("目前下載到 {}年{}月{}日，還有 {}天".format(ndate.year, ndate.month, ndate.day, daynum-_))
            time.sleep(5)

    filelist = os.listdir('C:\\Users\\GIGABYTE\\Downloads')
    for filename in filelist:
        if(number in filename):
            shutil.move('C:\\Users\\GIGABYTE\\Downloads\\' + filename, "C:\\Users\\GIGABYTE\\桌面\\VScode_Code\\天氣預測資料\\自動雨量站\\" + '{}({})'.format(name, number))
    driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > header > div.lightbox-tool-close').click()
    break
driver.quit()