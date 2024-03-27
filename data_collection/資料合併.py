import os
import openpyxl
import csv

frist_path = 'C:\\Users\\GIGABYTE\\桌面\\VScode_Code\\天氣預測資料'
frist_list = os.listdir(frist_path)
for case_name in frist_list:
    if("py" in case_name or "有人" in case_name or "氣象" in case_name or "雨量" in case_name):
        continue
    dir_path = os.path.join(frist_path, case_name)
    dir_list = os.listdir(dir_path)
    for dir_name in dir_list:
        if("py" in dir_name):
            continue
        file_path = os.path.join(dir_path, dir_name)
        file_list = os.listdir(file_path)
        workbook = openpyxl.load_workbook(os.path.join(file_path, dir_name+'.xlsx'))
        sheet = workbook.worksheets[0]
        for tmp in sheet.iter_rows(min_row=3, max_row=3, max_col=5, values_only=True):
            information = list(tmp)
        now_row = 3
        min_row = 3
        
        flag = True
        times = 0
        for file_name in file_list:
            if("xlsx" in file_name):
                continue
            csvfile = open(os.path.join(file_path, file_name), encoding='utf-8') # 開啟 CSV 檔案
            raw_data = csv.reader(csvfile) # 讀取 CSV 檔案
            data = list(raw_data) # 轉換成二維串列

            date = file_name.split('.')[0].split('-')[1:]
            date = '/'.join(date)
            
            if(flag):
                flag = False
                sheet.cell(row=1, column=6).value = '觀測年月日'
                sheet.cell(row=2, column=6).value = 'ObsDate'
                for index in range(len(data[0])):
                    if(index == 0):
                        data[0][0] = data[0][0].split('"')[1]
                    sheet.cell(row=1, column=7 + index).value = data[0][index]
                for index in range(len(data[1])):
                    sheet.cell(row=2, column=7 + index).value = data[1][index]
                workbook.save(os.path.join(file_path, dir_name+'.xlsx'))
            for csv_row in range(2, len(data)):
                sheet.cell(row=now_row, column=6).value = date
                for csv_col in range(len(data[0])):
                    sheet.cell(row=now_row, column=7 + csv_col).value = data[csv_row][csv_col]
                now_row += 1
            while(min_row < now_row):
                min_row += 1
                for index in range(len(information)):
                    sheet.cell(row=min_row, column=index+1).value = information[index]
            times += 1
            if(times%120 == 0):
                workbook.save(os.path.join(file_path, dir_name+'.xlsx'))
                print('現在測站 {} {} 整理到 {}'.format(case_name, dir_name, date))
                
        workbook.save(os.path.join(file_path, dir_name+'.xlsx'))
        workbook.close()